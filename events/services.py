import logging
from io import BytesIO

import openpyxl

from django.db import transaction
from django.utils import timezone

from .models import Event, Location

logger = logging.getLogger(__name__)


def export_events_to_xlsx(queryset):
    count = queryset.count()
    logger.info(f"Начало экспорта мероприятий. Количество к выгрузке: {count}")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Events"

    columns = [
        "Title",
        "Description",
        "Pub Date",
        "Start Date",
        "End Date",
        "Location Name",
        "Latitude",
        "Longitude",
        "Rating",
    ]
    sheet.append(columns)

    try:
        for event in queryset.select_related("location"):
            sheet.append(
                [
                    event.title,
                    event.description,
                    event.pub_date.replace(tzinfo=None) if event.pub_date else "",
                    event.start_date.replace(tzinfo=None) if event.start_date else "",
                    event.end_date.replace(tzinfo=None) if event.end_date else "",
                    event.location.name,
                    event.location.latitude,
                    event.location.longitude,
                    event.rating,
                ]
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        logger.info(f"Экспорт успешно завершен. Сформирован файл для {count} записей.")
        return output

    except Exception as e:
        # exception автоматически добавит traceback ошибки в лог
        logger.exception("Критическая ошибка при формировании XLSX файла")
        raise e


def import_events_from_xlsx(file_obj, user):
    from .tasks import update_single_location_weather

    logger.info("Запуск импорта XLSX.")

    try:
        workbook = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = workbook.active
    except Exception as e:
        logger.error(f"Не удалось открыть файл: {e}")
        raise ValueError("Ошибка формата файла") from e

    success_count = 0
    error_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            with transaction.atomic():
                title, desc, pub_d, start_d, end_d, loc_name, lat, lon, rating = row

                # Простейшая валидация данных (можно заменить на Serializer)
                if not title or not loc_name:
                    logger.warning(f"Строка {row_idx}: Пропуск. Отсутствует название или место.")
                    error_count += 1
                    continue

                location, _ = Location.objects.get_or_create(
                    name=loc_name, defaults={"latitude": lat, "longitude": lon}
                )

                def make_aware_if_naive(dt):
                    if dt and timezone.is_naive(dt):
                        return timezone.make_aware(dt)
                    return dt or timezone.now()

                Event.objects.create(
                    title=title,
                    description=desc,
                    pub_date=make_aware_if_naive(pub_d),
                    start_date=make_aware_if_naive(start_d),
                    end_date=make_aware_if_naive(end_d),
                    location=location,
                    author=user,
                    rating=rating or 0,
                    status="draft",
                )

                # Погода запрашивается только после успешного коммита строки
                transaction.on_commit(
                    lambda loc_id=location.id: update_single_location_weather.delay(loc_id)
                )

                success_count += 1

        except Exception as e:
            error_count += 1
            logger.error(f"Ошибка при обработке строки {row_idx}: {e}")
            # Продолжаем цикл, не прерывая весь импорт
            continue

    logger.info(f"Импорт завершен. Успешно: {success_count}, Ошибок: {error_count}")
    return success_count
